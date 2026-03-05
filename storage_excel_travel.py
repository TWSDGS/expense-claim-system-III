import os
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook

# Canonical column order for Travel
TRAVEL_COLUMNS: List[str] = [
    "id", 
    "status", 
    "filler_name", 
    "form_date", 
    "traveler_name", 
    "employee_no", "plan_code", 
    "purpose_desc", 
    "travel_route",
    "start_time", 
    "end_time", 
    "travel_days",
    "is_gov_car", "gov_car_no", 
    "is_taxi", 
    "is_private_car", "private_car_km", "private_car_no",
    "is_dispatch_car", 
    "is_hsr", 
    "is_airplane", 
    "is_other_transport", "other_transport_desc",
    "estimated_cost", 
    "expense_rows", 
    "total_amount",
    "handler_name", "project_manager_name", "dept_manager_name", "accountant_name",
    "attachments", 
    "created_at", 
    "updated_at", 
    "submitted_at"
]

TRAVEL_COLUMNS_ZH_MAP = {
    "id": "表單編號", 
    "status": "狀態", 
    "filler_name": "填表人", 
    "form_date": "填表日期", 
    "traveler_name": "出差人", 
    "plan_code": "計畫編號", 
    "purpose_desc": "出差事由", 
    "travel_route": "出差行程",
    "start_time": "出差起始時間", 
    "end_time": "出差結束時間", 
    "travel_days": "出差天數",
    "is_gov_car": "公務車", "gov_car_no": "公務車號", 
    "is_taxi": "計程車", 
    "is_private_car": "私車", "private_car_km": "私車公里數", "private_car_no": "私車車號",
    "is_dispatch_car": "派車", 
    "is_hsr": "高鐵", 
    "is_airplane": "飛機", 
    "is_other_transport": "其他交通", "other_transport_desc": "其他交通說明",
    "estimated_cost": "預估總花費", 
    "expense_rows": "出差明細(JSON)", 
    "total_amount": "總金額",
    "handler_name": "經手人", "project_manager_name": "計畫主持人", "dept_manager_name": "部門主管", "accountant_name": "會計",
    "attachments": "附件", 
    "created_at": "建立時間", 
    "updated_at": "更新時間", 
    "submitted_at": "送出時間"
}

def ensure_workbook(xlsx_path: str, sheet_name: str):
    os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)
    if os.path.exists(xlsx_path):
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(TRAVEL_COLUMNS)
            zh_cols = [TRAVEL_COLUMNS_ZH_MAP.get(c, c) for c in TRAVEL_COLUMNS]
            ws.append(zh_cols)
            wb.save(xlsx_path)
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(TRAVEL_COLUMNS)
    zh_cols = [TRAVEL_COLUMNS_ZH_MAP.get(c, c) for c in TRAVEL_COLUMNS]
    ws.append(zh_cols)
    wb.save(xlsx_path)

def _read_df(xlsx_path: str, sheet_name: str) -> pd.DataFrame:
    ensure_workbook(xlsx_path, sheet_name)
    try:
        # Try reading with MultiIndex (2 header rows)
        df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str, header=[0, 1])
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)
    except Exception:
        # Fallback to single header
        try:
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str)
        except Exception:
            ensure_workbook(xlsx_path, sheet_name)
            df = pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype=str)

    for c in TRAVEL_COLUMNS:
        if c not in df.columns:
            df[c] = ""
    df = df[TRAVEL_COLUMNS]
    return df.fillna("")

def load_all_travel(xlsx_path: str, draft_sheet: str, submit_sheet: str) -> pd.DataFrame:
    if not os.path.exists(xlsx_path):
        ensure_workbook(xlsx_path, draft_sheet)
        ensure_workbook(xlsx_path, submit_sheet)
        return pd.DataFrame(columns=TRAVEL_COLUMNS)
    
    dfs = []
    # Using openpyxl to check actual available sheets
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception:
        sheets = [draft_sheet, submit_sheet]
        
    for sn in set([draft_sheet, submit_sheet] + sheets):
        try:
           df = _read_df(xlsx_path, sn)
           dfs.append(df)
        except Exception:
           pass
           
    if not dfs:
        return pd.DataFrame(columns=TRAVEL_COLUMNS)
    df_all = pd.concat(dfs, ignore_index=True)
    if not df_all.empty and "id" in df_all.columns:
        df_all = df_all.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)
    return df_all

def _write_df_dual_header(xlsx_path: str, df: pd.DataFrame, sheet_name: str):
    zh_cols = [TRAVEL_COLUMNS_ZH_MAP.get(c, c) for c in TRAVEL_COLUMNS]
    df_zh = pd.DataFrame([zh_cols], columns=TRAVEL_COLUMNS)
    df_save = pd.concat([df_zh, df], ignore_index=True)
    
    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_save.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

def upsert_travel_record(xlsx_path: str, record: Dict, sheet_name: str):
    df = _read_df(xlsx_path, sheet_name)
    rid = str(record.get("id", "")).strip()
    if not rid:
        raise ValueError("record.id is required")

    row = {c: str(record.get(c, "")) for c in TRAVEL_COLUMNS}
    hit = df["id"] == rid
    if hit.any():
        df.loc[hit, TRAVEL_COLUMNS] = pd.DataFrame([row])[TRAVEL_COLUMNS].values
    else:
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    df = df.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)

    _write_df_dual_header(xlsx_path, df, sheet_name)

def delete_travel_record(xlsx_path: str, record_id: str, sheet_name: str):
    df = _read_df(xlsx_path, sheet_name)
    df = df[df["id"] != str(record_id)]
    df = df.sort_values("id", ascending=False, kind="mergesort").reset_index(drop=True)

    _write_df_dual_header(xlsx_path, df, sheet_name)

def cleanup_old_sheets(xlsx_path: str):
    if not os.path.exists(xlsx_path): return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)
        changed = False

        # Migrate data from old sheets to new sheets if new sheets don't exist or are empty
        mapping = {
            "出差申請單": "DomesticTrip",
            "出差草稿": "DomesticTrip_Draft"
        }
        for old_sn, new_sn in mapping.items():
            if old_sn in wb.sheetnames:
                if new_sn not in wb.sheetnames:
                    wb[old_sn].title = new_sn
                    changed = True
                else:
                    # new sheet exists. check if it has data.
                    # if new sheet is essentially empty (max_row <= 2), we can safely delete new sheet and rename old to new.
                    old_ws = wb[old_sn]
                    new_ws = wb[new_sn]
                    if new_ws.max_row <= 2:
                        del wb[new_sn]
                        old_ws.title = new_sn
                        changed = True
                    else:
                        # new sheet actually has data. we can't just overwrite it. just delete old to resolve conflict safely.
                        del wb[old_sn]
                        changed = True

        # Now remove any remaining obsolete sheets
        for old_sn in ["vouchers", "出差申請單", "出差草稿"]:
            if old_sn in wb.sheetnames:
                del wb[old_sn]
                changed = True
                
        # If we deleted sheets and there are no sheets left, create a default
        if not wb.sheetnames:
            wb.create_sheet("DomesticTrip")
            changed = True
        
        if changed:
            wb.save(xlsx_path)
    except Exception:
        pass
